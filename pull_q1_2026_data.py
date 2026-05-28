import os, pyodbc, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

conn_str = (
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={os.getenv('DB_SERVER')};"
    f"DATABASE={os.getenv('DB_DATABASE')};"
    f"UID={os.getenv('DB_USER')};"
    f"PWD={os.getenv('DB_PASSWORD')};"
    f"Encrypt=yes;TrustServerCertificate=no;"
)

print("Connecting to Azure SQL...")
conn = pyodbc.connect(conn_str, timeout=30)
print("Connected.")

QUERY = """
SELECT
    e.DEP_LASTDEPTHOSPITAL                          AS Hospital,
    e.DEP_LASTDEPTLOC                               AS Level_of_Care,
    e.ENC_PATCLASSBASE                              AS Patient_Class,
    e.TIME_HOSPADMISSION                            AS Admission_Time,
    e.TIME_HOSPDISCHARGE                            AS Discharge_Time,
    e.DISCHORDER_ORDERTIME                          AS Discharge_Order_Time,
    e.DISCHORDER_DISCHARGE                          AS DO_to_DC_min,
    e.DM_Complete_MedRec                            AS MedRec_Completion_Time,
    CASE
        WHEN e.DISCHORDER_ORDERTIME IS NOT NULL AND e.DM_Complete_MedRec IS NOT NULL
        THEN DATEDIFF(minute, e.DISCHORDER_ORDERTIME, e.DM_Complete_MedRec)
        ELSE NULL
    END                                             AS DO_to_MedRec_min,
    e.ENC_DISCHDISPO                                AS Discharge_Disposition
FROM DS_Encounters e
WHERE e.TIME_HOSPDISCHARGE >= '2026-01-01'
  AND e.TIME_HOSPDISCHARGE <  '2026-04-01'
  AND e.BEDDED = 'Y'
  AND e.PATIENT_IN_HOSPITAL_YN = 'N'
  AND e.TIME_HOSPDISCHARGE IS NOT NULL
ORDER BY e.DEP_LASTDEPTHOSPITAL, e.TIME_HOSPDISCHARGE
"""

print("Pulling Q1 2026 data...")
df = pd.read_sql(QUERY, conn)
conn.close()
print(f"Pulled {len(df):,} rows.")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Q1 2026 Discharge Data"

HEADERS = [
    "Hospital", "Level of Care", "Patient Class",
    "Admission Time", "Discharge Time", "Discharge Order Time",
    "DO→DC (min)", "Med Rec Completion Time", "DO→MedRec (min)",
    "Discharge Disposition"
]

# Header style
header_fill = PatternFill("solid", start_color="141533", end_color="141533")
header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[1].height = 30

# Write data
col_map = {
    "Hospital": "Hospital",
    "Level of Care": "Level_of_Care",
    "Patient Class": "Patient_Class",
    "Admission Time": "Admission_Time",
    "Discharge Time": "Discharge_Time",
    "Discharge Order Time": "Discharge_Order_Time",
    "DO→DC (min)": "DO_to_DC_min",
    "Med Rec Completion Time": "MedRec_Completion_Time",
    "DO→MedRec (min)": "DO_to_MedRec_min",
    "Discharge Disposition": "Discharge_Disposition"
}

data_font = Font(size=9, name="Arial")
data_align_center = Alignment(horizontal="center", vertical="center")
data_align_left = Alignment(horizontal="left", vertical="center")

for row_idx, row in enumerate(df.itertuples(index=False), 2):
    for col_idx, header in enumerate(HEADERS, 1):
        field = col_map[header]
        val = getattr(row, field)
        # Convert pandas NaT/nan to None
        if pd.isna(val) if not isinstance(val, str) else False:
            val = None
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = border
        if col_idx in (7, 9):  # numeric minute columns
            cell.alignment = data_align_center
            if val is not None:
                cell.number_format = '#,##0'
        elif col_idx in (4, 5, 6, 8):  # datetime columns
            cell.alignment = data_align_center
            if val is not None:
                cell.number_format = 'yyyy-mm-dd hh:mm'
        else:
            cell.alignment = data_align_left

    # Alternating row shading
    if row_idx % 2 == 0:
        row_fill = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill

# Column widths
col_widths = [28, 22, 16, 20, 20, 20, 14, 22, 16, 45]
for col_idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze header
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Q1_2026_Discharge_Data.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
